# -*- coding: utf-8 -*-
"""
gerar_excel.py — Toronto LTDA
Gera Toronto_LTDA_Relatorio.xlsx com 8 abas completas.
"""

import sqlite3
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

DB   = "toronto_ltda.db"
OUT  = "Toronto_LTDA_Relatorio.xlsx"
HOJE = "2026-06-02"

# ── estilos globais ───────────────────────────────────────────────────────────
FONT_NOME = "Arial"

def header_style(cell, bg="1F2937", fg="FFFFFF", bold=True, size=10):
    cell.font        = Font(name=FONT_NOME, bold=bold, color=fg, size=size)
    cell.fill        = PatternFill("solid", start_color=bg)
    cell.alignment   = Alignment(horizontal="center", vertical="center", wrap_text=True)

def kpi_style(cell, bg="111827", fg="F9A825", bold=True, size=12):
    cell.font        = Font(name=FONT_NOME, bold=bold, color=fg, size=size)
    cell.fill        = PatternFill("solid", start_color=bg)
    cell.alignment   = Alignment(horizontal="center", vertical="center")

def label_style(cell, bg="374151", fg="9CA3AF", size=8):
    cell.font        = Font(name=FONT_NOME, color=fg, size=size)
    cell.fill        = PatternFill("solid", start_color=bg)
    cell.alignment   = Alignment(horizontal="center", vertical="center")

def zebra(cell, row_idx):
    if row_idx % 2 == 0:
        cell.fill = PatternFill("solid", start_color="1F2937")
    else:
        cell.fill = PatternFill("solid", start_color="111827")
    cell.font      = Font(name=FONT_NOME, color="D1D5DB", size=9)
    cell.alignment = Alignment(vertical="center")

def thin_border():
    s = Side(style="thin", color="374151")
    return Border(left=s, right=s, top=s, bottom=s)

def set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

def autofit_rows(ws, height=18):
    for row in ws.iter_rows():
        ws.row_dimensions[row[0].row].height = height

FMT_BRL   = 'R$ #,##0.00'
FMT_BRL0  = 'R$ #,##0'
FMT_PCT   = '0.0%'
FMT_PCT1  = '0.0"%"'
FMT_INT   = '#,##0'
FMT_DATE  = 'DD/MM/YYYY'

GREEN  = "166534"
YELLOW = "92400E"
RED    = "991B1B"
BG_G   = "14532D"
BG_Y   = "78350F"
BG_R   = "7F1D1D"


def main():
    conn = sqlite3.connect(DB)
    wb   = Workbook()
    wb.remove(wb.active)   # remove sheet padrão

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 1 — CARTEIRA DE CLIENTES
    # ═══════════════════════════════════════════════════════════════════════════════

    df_cli = pd.read_sql("""
        SELECT
            c.id_cliente,
            COALESCE(c.cnpj_limpo, c.cnpj)                          AS cnpj,
            c.cnpj_valido,
            c.razao_social,
            c.nome_fantasia,
            c.segmento,
            c.cidade,
            c.uf,
            c.regiao,
            c.data_cadastro,
            v.nome                                                   AS vendedor,
            c.status,
            COUNT(DISTINCT fp.id_pedido)                             AS total_pedidos,
            ROUND(COALESCE(SUM(fp.valor_liquido), 0), 2)            AS faturamento_total,
            MAX(fp.data_pedido)                                      AS ultimo_pedido,
            ROUND(COALESCE(SUM(fp.valor_liquido) /
                  NULLIF(COUNT(DISTINCT fp.id_pedido), 0), 0), 2)   AS ticket_medio
        FROM dim_clientes c
        LEFT JOIN dim_vendedores v  ON c.id_vendedor  = v.id_vendedor
        LEFT JOIN fato_pedidos   fp ON c.id_cliente   = fp.id_cliente
                                    AND fp.status_pedido = 'Faturado'
        WHERE c.status = 'Ativo'
        GROUP BY c.id_cliente, c.cnpj_limpo, c.cnpj, c.cnpj_valido,
                 c.razao_social, c.nome_fantasia, c.segmento, c.cidade,
                 c.uf, c.regiao, c.data_cadastro, v.nome, c.status
        ORDER BY faturamento_total DESC
    """, conn)

    # Curva ABC
    fat_total = df_cli["faturamento_total"].sum()
    df_cli["fat_acum"] = df_cli["faturamento_total"].cumsum()
    df_cli["fat_acum_pct"] = df_cli["fat_acum"] / fat_total if fat_total > 0 else 0
    df_cli["curva_abc"] = df_cli["fat_acum_pct"].apply(
        lambda x: "A" if x <= 0.80 else ("B" if x <= 0.95 else "C")
    )
    df_cli["pct_faturamento"] = df_cli["faturamento_total"] / fat_total if fat_total > 0 else 0

    ws1 = wb.create_sheet("Carteira de Clientes")
    ws1.sheet_view.showGridLines = False
    ws1.tab_color = "C9A227"

    COLS1 = [
        ("ID",          "id_cliente",      6),
        ("CNPJ",        "cnpj",            18),
        ("CNPJ Válido", "cnpj_valido",     10),
        ("Razão Social","razao_social",    32),
        ("Fantasia",    "nome_fantasia",   22),
        ("Segmento",    "segmento",        14),
        ("Cidade",      "cidade",          16),
        ("UF",          "uf",               5),
        ("Região",      "regiao",          12),
        ("Cadastro",    "data_cadastro",   12),
        ("Vendedor",    "vendedor",        18),
        ("Status",      "status",           9),
        ("Pedidos",     "total_pedidos",    9),
        ("Faturamento", "faturamento_total",14),
        ("Últ. Pedido", "ultimo_pedido",   12),
        ("Ticket Médio","ticket_medio",    13),
        ("% Fat.",      "pct_faturamento",  9),
        ("ABC",         "curva_abc",        6),
    ]

    # Header
    ws1.row_dimensions[1].height = 30
    for c, (label, _, w) in enumerate(COLS1, 1):
        cell = ws1.cell(1, c, label)
        header_style(cell, bg="0F172A")
        set_col_width(ws1, c, w)

    # Dados
    for r, (_, row) in enumerate(df_cli.iterrows(), 2):
        ws1.row_dimensions[r].height = 17
        for c, (_, col, _) in enumerate(COLS1, 1):
            val = row[col]
            if pd.isna(val):
                val = ""
            cell = ws1.cell(r, c, val)
            zebra(cell, r)
            cell.border = thin_border()
            if col == "faturamento_total":
                cell.number_format = FMT_BRL0
            elif col == "ticket_medio":
                cell.number_format = FMT_BRL0
            elif col == "pct_faturamento":
                cell.number_format = FMT_PCT
            elif col == "data_cadastro" or col == "ultimo_pedido":
                cell.number_format = FMT_DATE
            elif col == "total_pedidos":
                cell.number_format = FMT_INT

    # Formatação condicional ABC
    last_row = len(df_cli) + 1
    abc_col  = get_column_letter(len(COLS1))
    ws1.conditional_formatting.add(
        f"{abc_col}2:{abc_col}{last_row}",
        CellIsRule(operator="equal", formula=['"A"'],
                   fill=PatternFill("solid", start_color="14532D"),
                   font=Font(color="86EFAC", bold=True))
    )
    ws1.conditional_formatting.add(
        f"{abc_col}2:{abc_col}{last_row}",
        CellIsRule(operator="equal", formula=['"B"'],
                   fill=PatternFill("solid", start_color="78350F"),
                   font=Font(color="FDE68A", bold=True))
    )
    ws1.conditional_formatting.add(
        f"{abc_col}2:{abc_col}{last_row}",
        CellIsRule(operator="equal", formula=['"C"'],
                   fill=PatternFill("solid", start_color="7F1D1D"),
                   font=Font(color="FCA5A5", bold=True))
    )

    freeze(ws1)
    print("  Aba 1 - Carteira de Clientes OK")

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 2 — GIRO FINANCEIRO
    # ═══════════════════════════════════════════════════════════════════════════════

    df_fin = pd.read_sql("""
        SELECT
            ff.id_titulo,
            ff.id_pedido,
            ff.id_cliente,
            COALESCE(c.nome_fantasia, c.razao_social) AS cliente,
            c.segmento,
            ff.data_emissao,
            ff.data_vencimento,
            ff.data_pagamento,
            ff.valor,
            ff.status_titulo,
            ff.dias_atraso
        FROM fato_financeiro ff
        JOIN dim_clientes c ON ff.id_cliente = c.id_cliente
        ORDER BY ff.data_vencimento DESC
    """, conn)

    ws2 = wb.create_sheet("Giro Financeiro")
    ws2.sheet_view.showGridLines = False
    ws2.tab_color = "3B82F6"

    # KPIs topo
    total_emit  = df_fin["valor"].sum()
    total_pago  = df_fin[df_fin["status_titulo"]=="Pago"]["valor"].sum()
    total_venc  = df_fin[df_fin["status_titulo"]=="Vencido"]["valor"].sum()
    total_aven  = df_fin[df_fin["status_titulo"]=="A Vencer"]["valor"].sum()
    taxa_inad   = len(df_fin[df_fin["status_titulo"]=="Vencido"]) / len(df_fin) * 100 if len(df_fin) else 0

    kpi_data = [
        ("Total Emitido",      total_emit,  FMT_BRL0, "1E3A5F"),
        ("Total Pago",         total_pago,  FMT_BRL0, "14532D"),
        ("Total Vencido",      total_venc,  FMT_BRL0, "7F1D1D"),
        ("A Vencer",           total_aven,  FMT_BRL0, "78350F"),
        ("Taxa Inadimplência", taxa_inad/100, FMT_PCT,"4A044E"),
    ]

    ws2.row_dimensions[1].height = 14
    ws2.row_dimensions[2].height = 12
    ws2.row_dimensions[3].height = 32
    ws2.row_dimensions[4].height = 12

    for i, (lbl, val, fmt, bg) in enumerate(kpi_data, 1):
        col_lbl = i * 3 - 2
        col_val = i * 3 - 1
        ws2.merge_cells(start_row=2, end_row=2, start_column=col_lbl, end_column=col_lbl+1)
        ws2.merge_cells(start_row=3, end_row=3, start_column=col_lbl, end_column=col_lbl+1)
        lc = ws2.cell(2, col_lbl, lbl)
        vc = ws2.cell(3, col_lbl, val)
        label_style(lc)
        kpi_style(vc)
        vc.fill   = PatternFill("solid", start_color=bg)
        vc.number_format = fmt
        for c2 in range(col_lbl, col_lbl+2):
            set_col_width(ws2, c2, 12)

    COLS2 = [
        ("ID Título",      "id_titulo",      9),
        ("ID Pedido",      "id_pedido",      9),
        ("ID Cliente",     "id_cliente",     9),
        ("Cliente",        "cliente",        28),
        ("Segmento",       "segmento",       14),
        ("Emissão",        "data_emissao",   12),
        ("Vencimento",     "data_vencimento",12),
        ("Pagamento",      "data_pagamento", 12),
        ("Valor",          "valor",          13),
        ("Status",         "status_titulo",  12),
        ("Dias Atraso",    "dias_atraso",    11),
    ]

    ROW_HDR2 = 6
    for c, (label, _, w) in enumerate(COLS2, 1):
        cell = ws2.cell(ROW_HDR2, c, label)
        header_style(cell, bg="0F172A")
        set_col_width(ws2, c, w)
    ws2.row_dimensions[ROW_HDR2].height = 28

    for r, (_, row) in enumerate(df_fin.iterrows(), ROW_HDR2+1):
        ws2.row_dimensions[r].height = 17
        for c, (_, col, _) in enumerate(COLS2, 1):
            val = row[col]
            if pd.isna(val): val = ""
            cell = ws2.cell(r, c, val)
            zebra(cell, r)
            cell.border = thin_border()
            if col == "valor":
                cell.number_format = FMT_BRL0
            elif col in ("data_emissao","data_vencimento","data_pagamento"):
                cell.number_format = FMT_DATE

    # Formatação condicional Status
    last2 = len(df_fin) + ROW_HDR2
    sc = get_column_letter(10)
    ws2.conditional_formatting.add(f"{sc}{ROW_HDR2+1}:{sc}{last2}",
        CellIsRule("equal",['"Pago"'],fill=PatternFill("solid",start_color="14532D"),font=Font(color="86EFAC")))
    ws2.conditional_formatting.add(f"{sc}{ROW_HDR2+1}:{sc}{last2}",
        CellIsRule("equal",['"Vencido"'],fill=PatternFill("solid",start_color="7F1D1D"),font=Font(color="FCA5A5")))
    ws2.conditional_formatting.add(f"{sc}{ROW_HDR2+1}:{sc}{last2}",
        CellIsRule("equal",['"A Vencer"'],fill=PatternFill("solid",start_color="78350F"),font=Font(color="FDE68A")))

    ws2.freeze_panes = f"A{ROW_HDR2+1}"
    print("  Aba 2 - Giro Financeiro OK")

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 3 — MARGEM POR PRODUTO
    # ═══════════════════════════════════════════════════════════════════════════════

    df_marg = pd.read_sql("""
        SELECT
            p.id_produto,
            p.sku,
            p.nome,
            p.categoria,
            p.pais_origem,
            p.fornecedor,
            p.custo_unit,
            p.preco_tabela,
            ROUND(p.preco_tabela - p.custo_unit, 2)                        AS margem_unit,
            ROUND(100.0*(p.preco_tabela - p.custo_unit)/p.preco_tabela, 1) AS margem_pct,
            COALESCE(SUM(fp.quantidade), 0)                                AS unidades_vendidas,
            ROUND(COALESCE(SUM(fp.valor_liquido), 0), 2)                   AS receita_liquida,
            ROUND(COALESCE(SUM(fp.quantidade),0)
                  * (p.preco_tabela - p.custo_unit), 2)                    AS lucro_bruto
        FROM dim_produtos p
        LEFT JOIN fato_pedidos fp ON p.id_produto = fp.id_produto
                                  AND fp.status_pedido = 'Faturado'
        GROUP BY p.id_produto, p.sku, p.nome, p.categoria,
                 p.pais_origem, p.fornecedor, p.custo_unit, p.preco_tabela
        ORDER BY lucro_bruto DESC
    """, conn)

    ws3 = wb.create_sheet("Margem por Produto")
    ws3.sheet_view.showGridLines = False
    ws3.tab_color = "10B981"

    # Bloco explicativo
    explicacao = [
        ("COMO A MARGEM É CALCULADA", "0F172A", "F9A825", 13),
        ("Margem Unitária (R$)",  "1F2937", "D1D5DB", 10),
        ("= Preço Tabela  −  Custo Unitário", "111827", "9CA3AF", 9),
        ("Margem % sobre Preço",  "1F2937", "D1D5DB", 10),
        ("= (Preço − Custo) ÷ Preço × 100", "111827", "9CA3AF", 9),
        ("Lucro Bruto Estimado (R$)", "1F2937", "D1D5DB", 10),
        ("= Unidades Vendidas × Margem Unitária", "111827", "9CA3AF", 9),
        ("","111827","9CA3AF",8),
    ]
    for i, (txt, bg, fg, sz) in enumerate(explicacao, 1):
        ws3.merge_cells(start_row=i, end_row=i, start_column=1, end_column=8)
        c = ws3.cell(i, 1, txt)
        c.font      = Font(name=FONT_NOME, color=fg or "9CA3AF", size=sz,
                           bold=(sz >= 10))
        c.fill      = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal="left", indent=1, vertical="center")
        ws3.row_dimensions[i].height = 16

    COLS3 = [
        ("ID",             "id_produto",      6),
        ("SKU",            "sku",            14),
        ("Produto",        "nome",            34),
        ("Categoria",      "categoria",       16),
        ("País Origem",    "pais_origem",     14),
        ("Fornecedor",     "fornecedor",      18),
        ("Custo Unit.",    "custo_unit",      11),
        ("Preço Tabela",   "preco_tabela",    11),
        ("Margem R$",      "margem_unit",     11),
        ("Margem %",       "margem_pct",      10),
        ("Un. Vendidas",   "unidades_vendidas",11),
        ("Receita Líq.",   "receita_liquida", 13),
        ("Lucro Bruto",    "lucro_bruto",     13),
    ]

    ROW_HDR3 = len(explicacao) + 1
    for c, (label, _, w) in enumerate(COLS3, 1):
        cell = ws3.cell(ROW_HDR3, c, label)
        header_style(cell, bg="0F172A")
        set_col_width(ws3, c, w)
    ws3.row_dimensions[ROW_HDR3].height = 28

    for r, (_, row) in enumerate(df_marg.iterrows(), ROW_HDR3+1):
        ws3.row_dimensions[r].height = 17
        for c, (_, col, _) in enumerate(COLS3, 1):
            val = row[col]
            if pd.isna(val): val = ""
            cell = ws3.cell(r, c, val)
            zebra(cell, r)
            cell.border = thin_border()
            if col in ("custo_unit","preco_tabela","margem_unit","receita_liquida","lucro_bruto"):
                cell.number_format = FMT_BRL0
            elif col == "margem_pct":
                cell.number_format = '0.0'
            elif col == "unidades_vendidas":
                cell.number_format = FMT_INT

    # Color scale na coluna margem_pct (verde claro → verde escuro)
    mp_col  = get_column_letter(10)
    last_r3 = ROW_HDR3 + len(df_marg)
    ws3.conditional_formatting.add(
        f"{mp_col}{ROW_HDR3+1}:{mp_col}{last_r3}",
        ColorScaleRule(start_type="min", start_color="FCA5A5",
                       mid_type="percentile", mid_value=50, mid_color="FDE68A",
                       end_type="max",   end_color="86EFAC")
    )

    ws3.freeze_panes = f"A{ROW_HDR3+1}"
    print("  Aba 3 - Margem por Produto OK")

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 4 — CROSS-SELL
    # ═══════════════════════════════════════════════════════════════════════════════

    # Pares de categorias comprados juntos no mesmo pedido
    df_cross_raw = pd.read_sql("""
        SELECT fp.id_pedido, p.categoria
        FROM fato_pedidos fp
        JOIN dim_produtos p ON fp.id_produto = p.id_produto
        WHERE fp.status_pedido = 'Faturado'
        GROUP BY fp.id_pedido, p.categoria
    """, conn)

    total_pedidos_fat = df_cross_raw["id_pedido"].nunique()

    # Pivot para pares
    from itertools import combinations
    pares = {}
    for pid, grp in df_cross_raw.groupby("id_pedido"):
        cats = sorted(grp["categoria"].unique())
        for a, b in combinations(cats, 2):
            k = (a, b)
            pares[k] = pares.get(k, 0) + 1

    df_pares = pd.DataFrame([
        {"Categoria A": k[0], "Categoria B": k[1],
         "Pedidos c/ ambas": v,
         "% dos pedidos": round(v / total_pedidos_fat * 100, 1)}
        for k, v in sorted(pares.items(), key=lambda x: -x[1])
    ])

    # Clientes por nº de categorias distintas
    df_cats_cli = pd.read_sql("""
        SELECT fp.id_cliente,
               COALESCE(c.nome_fantasia, c.razao_social) AS cliente,
               c.segmento,
               COUNT(DISTINCT p.categoria) AS categorias_distintas,
               GROUP_CONCAT(DISTINCT p.categoria) AS categorias
        FROM fato_pedidos fp
        JOIN dim_produtos p  ON fp.id_produto  = p.id_produto
        JOIN dim_clientes c  ON fp.id_cliente  = c.id_cliente
        WHERE fp.status_pedido = 'Faturado'
        GROUP BY fp.id_cliente, cliente, c.segmento
        ORDER BY categorias_distintas DESC
    """, conn)
    df_cats_cli["oportunidade"] = df_cats_cli["categorias_distintas"].apply(
        lambda x: "Monoproduto — Cross-sell!" if x == 1 else
                  ("Baixo mix" if x == 2 else "Bom mix")
    )

    ws4 = wb.create_sheet("Cross-sell")
    ws4.sheet_view.showGridLines = False
    ws4.tab_color = "8B5CF6"

    # Título
    ws4.merge_cells("A1:G1")
    t = ws4.cell(1, 1, "ANALISE DE CROSS-SELL — Pares de Categorias Comprados Juntos")
    header_style(t, bg="1E1B4B", fg="A78BFA", size=11)
    ws4.row_dimensions[1].height = 28

    # Tabela pares
    HPARES = ["Categoria A","Categoria B","Pedidos c/ Ambas","% dos Pedidos"]
    for c, h in enumerate(HPARES, 1):
        cell = ws4.cell(3, c, h)
        header_style(cell, bg="0F172A")
        set_col_width(ws4, c, 20 if c <= 2 else 16)
    ws4.row_dimensions[3].height = 26

    for r, (_, row) in enumerate(df_pares.iterrows(), 4):
        ws4.row_dimensions[r].height = 16
        for c, col in enumerate(["Categoria A","Categoria B","Pedidos c/ ambas","% dos pedidos"], 1):
            cell = ws4.cell(r, c, row[col])
            zebra(cell, r)
            cell.border = thin_border()
            if col == "% dos pedidos":
                cell.number_format = '0.0"%"'

    # Separador
    sep_row = 4 + len(df_pares) + 1
    ws4.merge_cells(start_row=sep_row, end_row=sep_row, start_column=1, end_column=8)
    sep = ws4.cell(sep_row, 1, "OPORTUNIDADE DE CROSS-SELL POR CLIENTE")
    header_style(sep, bg="1E1B4B", fg="A78BFA", size=10)
    ws4.row_dimensions[sep_row].height = 22

    HCLI = ["ID Cliente","Cliente","Segmento","Categorias Distintas","Categorias Compradas","Oportunidade"]
    for c, h in enumerate(HCLI, 1):
        cell = ws4.cell(sep_row+1, c, h)
        header_style(cell, bg="0F172A")
        ws = [8, 28, 14, 16, 42, 22]
        set_col_width(ws4, c, ws[c-1])

    for r, (_, row) in enumerate(df_cats_cli.iterrows(), sep_row+2):
        ws4.row_dimensions[r].height = 16
        for c, col in enumerate(["id_cliente","cliente","segmento","categorias_distintas","categorias","oportunidade"], 1):
            cell = ws4.cell(r, c, row[col])
            zebra(cell, r)
            cell.border = thin_border()

    # Cor na coluna oportunidade
    op_col = get_column_letter(6)
    last_r4 = sep_row + 1 + len(df_cats_cli)
    ws4.conditional_formatting.add(f"{op_col}{sep_row+2}:{op_col}{last_r4}",
        CellIsRule("equal",['"Monoproduto — Cross-sell!"'],
                   fill=PatternFill("solid",start_color="7F1D1D"),
                   font=Font(color="FCA5A5")))
    ws4.conditional_formatting.add(f"{op_col}{sep_row+2}:{op_col}{last_r4}",
        CellIsRule("equal",['"Baixo mix"'],
                   fill=PatternFill("solid",start_color="78350F"),
                   font=Font(color="FDE68A")))
    ws4.conditional_formatting.add(f"{op_col}{sep_row+2}:{op_col}{last_r4}",
        CellIsRule("equal",['"Bom mix"'],
                   fill=PatternFill("solid",start_color="14532D"),
                   font=Font(color="86EFAC")))

    ws4.freeze_panes = "A4"
    print("  Aba 4 - Cross-sell OK")

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 5 — TAXA DE CONVERSÃO
    # ═══════════════════════════════════════════════════════════════════════════════

    df_conv = pd.read_sql("""
        SELECT
            v.nome                                               AS vendedor,
            COUNT(DISTINCT fp.id_pedido)                        AS total_pedidos,
            COUNT(DISTINCT CASE WHEN fp.status_pedido='Faturado'   THEN fp.id_pedido END) AS faturados,
            COUNT(DISTINCT CASE WHEN fp.status_pedido='Cancelado'  THEN fp.id_pedido END) AS cancelados,
            COUNT(DISTINCT CASE WHEN fp.status_pedido='Devolvido'  THEN fp.id_pedido END) AS devolvidos,
            ROUND(100.0 * COUNT(DISTINCT CASE WHEN fp.status_pedido='Faturado' THEN fp.id_pedido END)
                  / NULLIF(COUNT(DISTINCT fp.id_pedido), 0), 1)  AS taxa_conversao
        FROM fato_pedidos fp
        JOIN dim_vendedores v ON fp.id_vendedor = v.id_vendedor
        GROUP BY v.id_vendedor, v.nome
        ORDER BY taxa_conversao DESC
    """, conn)

    df_conv_mes = pd.read_sql("""
        SELECT
            strftime('%Y-%m', data_pedido)                            AS mes,
            COUNT(DISTINCT id_pedido)                                 AS total,
            COUNT(DISTINCT CASE WHEN status_pedido='Faturado'   THEN id_pedido END) AS faturados,
            COUNT(DISTINCT CASE WHEN status_pedido='Cancelado'  THEN id_pedido END) AS cancelados,
            COUNT(DISTINCT CASE WHEN status_pedido='Devolvido'  THEN id_pedido END) AS devolvidos,
            ROUND(100.0 * COUNT(DISTINCT CASE WHEN status_pedido='Faturado' THEN id_pedido END)
                  / NULLIF(COUNT(DISTINCT id_pedido), 0), 1)         AS taxa_conversao
        FROM fato_pedidos
        GROUP BY mes ORDER BY mes
    """, conn)

    ws5 = wb.create_sheet("Taxa de Conversao")
    ws5.sheet_view.showGridLines = False
    ws5.tab_color = "F59E0B"

    # Tabela por vendedor
    ws5.merge_cells("A1:F1")
    t5 = ws5.cell(1, 1, "TAXA DE CONVERSAO DE ORCAMENTOS EM PEDIDOS — POR VENDEDOR")
    header_style(t5, bg="1C1917", fg="FCD34D", size=11)
    ws5.row_dimensions[1].height = 28

    HCONV = ["Vendedor","Total Pedidos","Faturados","Cancelados","Devolvidos","Taxa Conversao %"]
    wconv = [22, 13, 11, 11, 11, 16]
    for c, (h, w) in enumerate(zip(HCONV, wconv), 1):
        cell = ws5.cell(2, c, h)
        header_style(cell, bg="0F172A")
        set_col_width(ws5, c, w)
    ws5.row_dimensions[2].height = 26

    for r, (_, row) in enumerate(df_conv.iterrows(), 3):
        ws5.row_dimensions[r].height = 17
        for c, col in enumerate(["vendedor","total_pedidos","faturados","cancelados","devolvidos","taxa_conversao"], 1):
            cell = ws5.cell(r, c, row[col])
            zebra(cell, r)
            cell.border = thin_border()
            if col == "taxa_conversao":
                cell.number_format = '0.0"%"'
            else:
                cell.number_format = FMT_INT

    # Condicional taxa
    tc_col  = get_column_letter(6)
    last_r5 = 2 + len(df_conv)
    ws5.conditional_formatting.add(f"{tc_col}3:{tc_col}{last_r5}",
        CellIsRule("greaterThanOrEqual",["85"],fill=PatternFill("solid",start_color="14532D"),font=Font(color="86EFAC",bold=True)))
    ws5.conditional_formatting.add(f"{tc_col}3:{tc_col}{last_r5}",
        CellIsRule("between",["70","85"],fill=PatternFill("solid",start_color="78350F"),font=Font(color="FDE68A")))
    ws5.conditional_formatting.add(f"{tc_col}3:{tc_col}{last_r5}",
        CellIsRule("lessThan",["70"],fill=PatternFill("solid",start_color="7F1D1D"),font=Font(color="FCA5A5")))

    # Tabela por mês
    sep5 = last_r5 + 2
    ws5.merge_cells(start_row=sep5, end_row=sep5, start_column=1, end_column=6)
    t5b = ws5.cell(sep5, 1, "TAXA DE CONVERSAO MENSAL")
    header_style(t5b, bg="1C1917", fg="FCD34D", size=10)
    ws5.row_dimensions[sep5].height = 22

    for c, (h, w) in enumerate(zip(["Mês","Total","Faturados","Cancelados","Devolvidos","Taxa %"], wconv), 1):
        cell = ws5.cell(sep5+1, c, h)
        header_style(cell, bg="0F172A")

    for r, (_, row) in enumerate(df_conv_mes.iterrows(), sep5+2):
        ws5.row_dimensions[r].height = 16
        for c, col in enumerate(["mes","total","faturados","cancelados","devolvidos","taxa_conversao"], 1):
            cell = ws5.cell(r, c, row[col])
            zebra(cell, r)
            cell.border = thin_border()
            if col == "taxa_conversao": cell.number_format = '0.0"%"'

    ws5.freeze_panes = "A3"
    print("  Aba 5 - Taxa de Conversao OK")

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 6 — PEDIDOS NÃO FATURADOS
    # ═══════════════════════════════════════════════════════════════════════════════

    df_nfat = pd.read_sql("""
        SELECT
            fp.id_pedido,
            fp.data_pedido,
            COALESCE(c.nome_fantasia, c.razao_social) AS cliente,
            c.segmento,
            v.nome                                     AS vendedor,
            fp.status_pedido,
            COUNT(fp.id_item)                          AS qtd_itens,
            ROUND(SUM(fp.valor_bruto), 2)              AS valor_bruto_perdido,
            ROUND(SUM(fp.valor_liquido), 2)            AS valor_liquido_perdido
        FROM fato_pedidos fp
        JOIN dim_clientes  c ON fp.id_cliente  = c.id_cliente
        JOIN dim_vendedores v ON fp.id_vendedor = v.id_vendedor
        WHERE fp.status_pedido IN ('Cancelado','Devolvido')
        GROUP BY fp.id_pedido, fp.data_pedido, cliente, c.segmento,
                 v.nome, fp.status_pedido
        ORDER BY fp.data_pedido DESC
    """, conn)

    ws6 = wb.create_sheet("Pedidos nao Faturados")
    ws6.sheet_view.showGridLines = False
    ws6.tab_color = "EF4444"

    tot_canc  = len(df_nfat[df_nfat.status_pedido=="Cancelado"])
    tot_dev   = len(df_nfat[df_nfat.status_pedido=="Devolvido"])
    val_perd  = df_nfat["valor_liquido_perdido"].sum()

    kpis6 = [
        ("Total Cancelados",  tot_canc,  FMT_INT,  "7F1D1D"),
        ("Total Devolvidos",  tot_dev,   FMT_INT,  "78350F"),
        ("Valor Total Perdido", val_perd, FMT_BRL0, "4C1D95"),
    ]
    ws6.row_dimensions[1].height = 14
    ws6.row_dimensions[2].height = 12
    ws6.row_dimensions[3].height = 32
    ws6.row_dimensions[4].height = 12

    for i, (lbl, val, fmt, bg) in enumerate(kpis6, 1):
        col = i * 3 - 2
        ws6.merge_cells(start_row=2, end_row=2, start_column=col, end_column=col+1)
        ws6.merge_cells(start_row=3, end_row=3, start_column=col, end_column=col+1)
        lc = ws6.cell(2, col, lbl); label_style(lc)
        vc = ws6.cell(3, col, val); kpi_style(vc)
        vc.fill = PatternFill("solid", start_color=bg)
        vc.number_format = fmt
        for c2 in range(col, col+2): set_col_width(ws6, c2, 14)

    COLS6 = [
        ("ID Pedido",  "id_pedido",            9),
        ("Data",       "data_pedido",          12),
        ("Cliente",    "cliente",              28),
        ("Segmento",   "segmento",             14),
        ("Vendedor",   "vendedor",             18),
        ("Status",     "status_pedido",        12),
        ("Itens",      "qtd_itens",             7),
        ("Vlr. Bruto", "valor_bruto_perdido",  13),
        ("Vlr. Líq.",  "valor_liquido_perdido",13),
    ]

    ROW_H6 = 6
    for c, (label, _, w) in enumerate(COLS6, 1):
        cell = ws6.cell(ROW_H6, c, label)
        header_style(cell, bg="0F172A")
        set_col_width(ws6, c, w)
    ws6.row_dimensions[ROW_H6].height = 28

    for r, (_, row) in enumerate(df_nfat.iterrows(), ROW_H6+1):
        ws6.row_dimensions[r].height = 17
        for c, (_, col, _) in enumerate(COLS6, 1):
            val = row[col]
            if pd.isna(val): val = ""
            cell = ws6.cell(r, c, val)
            zebra(cell, r)
            cell.border = thin_border()
            if col in ("valor_bruto_perdido","valor_liquido_perdido"):
                cell.number_format = FMT_BRL0
            elif col == "data_pedido":
                cell.number_format = FMT_DATE

    st_col  = get_column_letter(6)
    last_r6 = ROW_H6 + len(df_nfat)
    ws6.conditional_formatting.add(f"{st_col}{ROW_H6+1}:{st_col}{last_r6}",
        CellIsRule("equal",['"Cancelado"'],fill=PatternFill("solid",start_color="7F1D1D"),font=Font(color="FCA5A5")))
    ws6.conditional_formatting.add(f"{st_col}{ROW_H6+1}:{st_col}{last_r6}",
        CellIsRule("equal",['"Devolvido"'],fill=PatternFill("solid",start_color="78350F"),font=Font(color="FDE68A")))

    ws6.freeze_panes = f"A{ROW_H6+1}"
    print("  Aba 6 - Pedidos nao Faturados OK")

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 7 — TOP 10 PRODUTOS
    # ═══════════════════════════════════════════════════════════════════════════════

    df_top = pd.read_sql("""
        SELECT
            p.nome,
            p.categoria,
            p.fornecedor,
            ROUND(SUM(fp.valor_liquido), 2)             AS receita_liquida,
            SUM(fp.quantidade)                          AS unidades_vendidas,
            ROUND(AVG(fp.valor_liquido), 2)             AS ticket_medio_item
        FROM fato_pedidos fp
        JOIN dim_produtos p ON fp.id_produto = p.id_produto
        WHERE fp.status_pedido = 'Faturado'
        GROUP BY p.id_produto, p.nome, p.categoria, p.fornecedor
        ORDER BY receita_liquida DESC
        LIMIT 10
    """, conn)

    fat_total_top = pd.read_sql(
        "SELECT SUM(valor_liquido) AS t FROM fato_pedidos WHERE status_pedido='Faturado'", conn
    ).iloc[0]["t"]

    df_top["ranking"]          = range(1, len(df_top)+1)
    df_top["pct_faturamento"]  = df_top["receita_liquida"] / fat_total_top
    df_top["receita_acum"]     = df_top["receita_liquida"].cumsum()
    df_top["pct_acum"]         = df_top["receita_acum"] / fat_total_top

    ws7 = wb.create_sheet("Top 10 Produtos")
    ws7.sheet_view.showGridLines = False
    ws7.tab_color = "F97316"

    ws7.merge_cells("A1:I1")
    t7 = ws7.cell(1, 1, "TOP 10 PRODUTOS — REPRESENTATIVIDADE NO FATURAMENTO TOTAL")
    header_style(t7, bg="1C0A00", fg="FB923C", size=12)
    ws7.row_dimensions[1].height = 30

    COLS7 = [
        ("Rank",          "ranking",         6),
        ("Produto",        "nome",           36),
        ("Categoria",      "categoria",      16),
        ("Fornecedor",     "fornecedor",     18),
        ("Receita Líq.",   "receita_liquida",14),
        ("Un. Vendidas",   "unidades_vendidas",12),
        ("Ticket Médio",   "ticket_medio_item",13),
        ("% do Fat.",      "pct_faturamento", 10),
        ("% Acum.",        "pct_acum",         9),
    ]

    for c, (label, _, w) in enumerate(COLS7, 1):
        cell = ws7.cell(2, c, label)
        header_style(cell, bg="0F172A")
        set_col_width(ws7, c, w)
    ws7.row_dimensions[2].height = 26

    RANK_COLORS = ["C9A227","B8961F","A78A1A","967F16","857312","746712","635B0E","52500A","413F07","2F2E04"]
    for r, (_, row) in enumerate(df_top.iterrows(), 3):
        ws7.row_dimensions[r].height = 22
        bg_rank = RANK_COLORS[r-3]
        for c, (_, col, _) in enumerate(COLS7, 1):
            val = row[col]
            if pd.isna(val): val = ""
            cell = ws7.cell(r, c, val)
            if c == 1:
                cell.font  = Font(name=FONT_NOME, bold=True, color="0F172A", size=11)
                cell.fill  = PatternFill("solid", start_color=bg_rank)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                zebra(cell, r)
            cell.border = thin_border()
            if col in ("receita_liquida","ticket_medio_item"):
                cell.number_format = FMT_BRL0
            elif col in ("pct_faturamento","pct_acum"):
                cell.number_format = FMT_PCT
            elif col == "unidades_vendidas":
                cell.number_format = FMT_INT

    ws7.freeze_panes = "A3"
    print("  Aba 7 - Top 10 Produtos OK")

    # ═══════════════════════════════════════════════════════════════════════════════
    # ABA 8 — PRODUTOS EM QUEDA
    # ═══════════════════════════════════════════════════════════════════════════════

    # Períodos: últimos 6 meses vs. 6 meses anteriores
    df_pq = pd.read_sql("""
        WITH periodos AS (
            SELECT
                p.id_produto,
                p.nome,
                p.categoria,
                p.pais_origem,
                p.fornecedor,
                ROUND(SUM(CASE WHEN fp.data_pedido BETWEEN '2025-06-01' AND '2025-11-30'
                          THEN fp.valor_liquido ELSE 0 END), 2) AS fat_anterior,
                ROUND(SUM(CASE WHEN fp.data_pedido BETWEEN '2025-12-01' AND '2026-05-31'
                          THEN fp.valor_liquido ELSE 0 END), 2) AS fat_recente
            FROM dim_produtos p
            LEFT JOIN fato_pedidos fp ON p.id_produto = fp.id_produto
                                     AND fp.status_pedido = 'Faturado'
            GROUP BY p.id_produto, p.nome, p.categoria, p.pais_origem, p.fornecedor
        )
        SELECT *,
               ROUND(fat_recente - fat_anterior, 2) AS variacao_abs,
               CASE WHEN fat_anterior > 0
                    THEN ROUND((fat_recente - fat_anterior) / fat_anterior * 100, 1)
                    ELSE NULL END AS variacao_pct
        FROM periodos
        WHERE fat_anterior > 0 OR fat_recente > 0
        ORDER BY variacao_pct ASC
    """, conn)

    def gerar_insight(row):
        pct = row["variacao_pct"]
        if pd.isna(pct) or pct >= -5:
            return ""
        cat = str(row.get("categoria",""))
        pais = str(row.get("pais_origem",""))
        forn = str(row.get("fornecedor",""))
        causas = []
        if pct < -30:
            causas.append("Queda acentuada — possível ruptura de estoque ou perda de cliente-chave")
        if "Destilado" in cat:
            causas.append("Categoria sensível a sazonalidade (picos em fim de ano)")
        if "Espumante" in cat:
            causas.append("Espumantes têm forte sazonalidade — análise recomendada fora de dez/jan")
        if pais in ("Argentina","Chile"):
            causas.append("Câmbio e importação afetam custo — revisar precificação")
        if not causas:
            causas.append("Revisar mix ativo com vendedor responsável; possível perda de carteira")
        return "; ".join(causas)

    df_pq["status"] = df_pq["variacao_pct"].apply(
        lambda x: "Em Queda" if pd.notna(x) and x < -5
        else ("Crescimento" if pd.notna(x) and x >= 5 else "Estavel")
    )
    df_pq["possivel_causa"] = df_pq.apply(gerar_insight, axis=1)
    df_pq = df_pq.sort_values("variacao_pct", ascending=True, na_position="last")

    ws8 = wb.create_sheet("Produtos em Queda")
    ws8.sheet_view.showGridLines = False
    ws8.tab_color = "DC2626"

    ws8.merge_cells("A1:J1")
    t8 = ws8.cell(1, 1, "PRODUTOS EM QUEDA — Comparativo 6 Meses (Jun–Nov 2025 vs Dez 2025–Mai 2026)")
    header_style(t8, bg="450A0A", fg="FCA5A5", size=11)
    ws8.row_dimensions[1].height = 28

    COLS8 = [
        ("Produto",     "nome",         32),
        ("Categoria",   "categoria",    16),
        ("País",        "pais_origem",  12),
        ("Fornecedor",  "fornecedor",   18),
        ("Fat. Anterior","fat_anterior",13),
        ("Fat. Recente","fat_recente",  13),
        ("Var. R$",     "variacao_abs", 12),
        ("Var. %",      "variacao_pct", 10),
        ("Status",      "status",       12),
        ("Possível Causa","possivel_causa",44),
    ]

    for c, (label, _, w) in enumerate(COLS8, 1):
        cell = ws8.cell(2, c, label)
        header_style(cell, bg="0F172A")
        set_col_width(ws8, c, w)
    ws8.row_dimensions[2].height = 26

    for r, (_, row) in enumerate(df_pq.iterrows(), 3):
        ws8.row_dimensions[r].height = 18
        for c, (_, col, _) in enumerate(COLS8, 1):
            val = row[col]
            if pd.isna(val): val = ""
            cell = ws8.cell(r, c, val)
            zebra(cell, r)
            cell.border = thin_border()
            if col in ("fat_anterior","fat_recente","variacao_abs"):
                cell.number_format = FMT_BRL0
            elif col == "variacao_pct":
                cell.number_format = '0.0"%"'
            if col == "possivel_causa":
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    st_col8  = get_column_letter(9)
    last_r8  = 2 + len(df_pq)
    ws8.conditional_formatting.add(f"{st_col8}3:{st_col8}{last_r8}",
        CellIsRule("equal",['"Em Queda"'],fill=PatternFill("solid",start_color="7F1D1D"),font=Font(color="FCA5A5",bold=True)))
    ws8.conditional_formatting.add(f"{st_col8}3:{st_col8}{last_r8}",
        CellIsRule("equal",['"Crescimento"'],fill=PatternFill("solid",start_color="14532D"),font=Font(color="86EFAC",bold=True)))
    ws8.conditional_formatting.add(f"{st_col8}3:{st_col8}{last_r8}",
        CellIsRule("equal",['"Estavel"'],fill=PatternFill("solid",start_color="1E3A5F"),font=Font(color="93C5FD")))

    # Color scale na variação %
    vp_col = get_column_letter(8)
    ws8.conditional_formatting.add(f"{vp_col}3:{vp_col}{last_r8}",
        ColorScaleRule(start_type="min", start_color="EF4444",
                       mid_type="num",   mid_value=0,   mid_color="FFFFFF",
                       end_type="max",   end_color="22C55E"))

    ws8.freeze_panes = "A3"
    print("  Aba 8 - Produtos em Queda OK")

    # ── salvar ────────────────────────────────────────────────────────────────────
    conn.close()
    wb.save(OUT)
    print(f"\nPlanilha salva: {OUT}")


if __name__ == "__main__":
    main()
